from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.result import CollectStats, ContentItem


VIEW_COLUMNS = [
    "keyword",
    "title",
    "author_name",
    "like_count",
    "comment_count",
    "publish_time",
    "content_length",
    "image_count",
    "field_completeness",
    "quality_score",
    "lead_score",
    "parse_status",
    "detail_error",
    "risk_level",
    "url",
]

RAW_COLUMNS = [
    "source_platform",
    "source_type",
    "source_id",
    "url",
    "keyword",
    "task_id",
    "title",
    "snippet",
    "cover_url",
    "author_name",
    "author_id",
    "author_home_url",
    "like_count",
    "content_text",
    "image_urls",
    "image_count",
    "publish_time",
    "comment_count",
    "share_count",
    "collect_count",
    "tags",
    "parse_stage",
    "parse_status",
    "detail_attempted",
    "detail_error",
    "drop_reason",
    "field_completeness",
    "engagement_score",
    "quality_score",
    "lead_score",
    "risk_level",
    "risk_reason",
    "has_contact_hint",
    "content_length",
    "is_detail_complete",
    "collected_at",
    "updated_at",
]


def _stringify(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(x) for x in value)
    return value


def _write_header(ws, columns: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")


def _write_rows(ws, columns: list[str], rows: list[dict]) -> None:
    for r_idx, item in enumerate(rows, 2):
        for c_idx, col in enumerate(columns, 1):
            cell = ws.cell(r_idx, c_idx, _stringify(item.get(col)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _auto_fit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def _color_status(ws, columns: list[str], status_col_name: str = "parse_status") -> None:
    if status_col_name not in columns:
        return

    idx = columns.index(status_col_name) + 1
    fills = {
        "detail_success": PatternFill("solid", fgColor="C6EFCE"),
        "detail_failed": PatternFill("solid", fgColor="FFC7CE"),
        "list_only": PatternFill("solid", fgColor="FFEB9C"),
        "dropped": PatternFill("solid", fgColor="D9D9D9"),
    }

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, idx)
        fill = fills.get(cell.value)
        if fill:
            cell.fill = fill


def _build_summary_rows(stats: dict, items: list[dict]) -> list[tuple[str, float | int]]:
    total = len(items)
    detail_complete_count = sum(1 for x in items if x.get("is_detail_complete"))
    avg_like = round(sum((x.get("like_count") or 0) for x in items) / total, 2) if total else 0
    comment_items = [x for x in items if x.get("comment_count") is not None]
    avg_comment = (
        round(sum((x.get("comment_count") or 0) for x in comment_items) / len(comment_items), 2)
        if comment_items
        else 0
    )

    return [
        ("total_items", total),
        ("discovered", stats.get("discovered", 0)),
        ("list_success", stats.get("list_success", 0)),
        ("detail_attempted", stats.get("detail_attempted", 0)),
        ("detail_success", stats.get("detail_success", 0)),
        ("detail_failed", stats.get("detail_failed", 0)),
        ("dropped", stats.get("dropped", 0)),
        ("detail_complete_count", detail_complete_count),
        ("detail_complete_ratio", round(detail_complete_count / total, 4) if total else 0),
        ("avg_like_count", avg_like),
        ("avg_comment_count", avg_comment),
    ]


def export_sample_to_excel(
    items: list[ContentItem],
    keyword: str,
    platform: str,
    sample_size: int,
    export_dir: str,
    stats: CollectStats | None = None,
) -> str:
    sample_items = items[:sample_size]

    target_dir = Path(export_dir)
    target_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = "".join(ch if ch.isalnum() else "_" for ch in keyword).strip("_") or "keyword"
    file_name = f"sample_{platform}_{safe_keyword}_{timestamp}.xlsx"
    file_path = target_dir / file_name

    rows = [item.model_dump(mode="json") for item in sample_items]
    summary_stats = stats.model_dump() if stats else {}

    wb = Workbook()

    ws_view = wb.active
    ws_view.title = "items_view"
    _write_header(ws_view, VIEW_COLUMNS)
    _write_rows(ws_view, VIEW_COLUMNS, rows)
    _color_status(ws_view, VIEW_COLUMNS)
    ws_view.freeze_panes = "A2"
    _auto_fit(ws_view)

    ws_raw = wb.create_sheet("items_raw")
    _write_header(ws_raw, RAW_COLUMNS)
    _write_rows(ws_raw, RAW_COLUMNS, rows)
    _color_status(ws_raw, RAW_COLUMNS)
    ws_raw.freeze_panes = "A2"
    _auto_fit(ws_raw)

    ws_sum = wb.create_sheet("summary")
    ws_sum["A1"] = "metric"
    ws_sum["B1"] = "value"
    ws_sum["A1"].font = Font(bold=True)
    ws_sum["B1"].font = Font(bold=True)
    for idx, (k, v) in enumerate(_build_summary_rows(summary_stats, rows), start=2):
        ws_sum.cell(row=idx, column=1, value=k)
        ws_sum.cell(row=idx, column=2, value=v)
    _auto_fit(ws_sum)

    wb.save(file_path)
    return str(file_path.resolve())
